from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.utils import timezone
from django.db import transaction
from django.db.models import Count, Avg
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from openpyxl import load_workbook
from datetime import datetime
import json
from .models import (
    Materia, Mensaje, PerfilEstudiante, PerfilDocente, PerfilUsuario, 
    Carrera, Matricula, Periodo, Tarea, EntregaTarea, Certificado, Calificacion,
    Planificacion, Parcial, Unidad, ActividadPlanificada
)


# ===== DECORATOR: Protección por rol =====
def rol_requerido(*roles_permitidos):
    """Decorator que valida si el usuario tiene uno de los roles requeridos"""
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            
            try:
                perfil = PerfilUsuario.objects.get(user=request.user)
                if perfil.rol not in roles_permitidos:
                    messages.error(request, "❌ No tienes permisos para acceder a esta sección")
                    return redirect('index')
            except PerfilUsuario.DoesNotExist:
                messages.error(request, "❌ Tu perfil no está configurado")
                return redirect('index')
            
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


# ===== VISTAS PÚBLICAS =====

def index(request):
    """Página principal"""
    return render(request, 'plataforma/index.html')

def aulas_virtuales(request):
    """Aulas virtuales - redirige a login o al dashboard"""
    if request.user.is_authenticated:
        return redirect('dashboard_estudiante')
    return redirect('login')

def portal_estudiantil(request):
    """Portal estudiantil - lista de estudiantes con su carrera"""
    estudiantes = PerfilEstudiante.objects.select_related('user', 'carrera').filter(activo=True)
    return render(request, 'plataforma/portal_estudiantil.html', {'estudiantes': estudiantes})

def biblioteca(request):
    """Biblioteca digital"""
    return render(request, 'plataforma/biblioteca.html')

def admisiones(request):
    """Admisiones"""
    return render(request, 'plataforma/admisiones.html')

def contacto(request):
    """Formulario de contacto"""
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        email = request.POST.get('email')
        asunto = request.POST.get('asunto')
        mensaje_texto = request.POST.get('mensaje')

        Mensaje.objects.create(
            nombre=nombre,
            email=email,
            asunto=asunto,
            mensaje=mensaje_texto
        )

        return redirect('contacto_exito')

    return render(request, 'plataforma/contacto.html')

def contacto_exito(request):
    """Página de éxito del contacto"""
    return render(request, 'plataforma/contacto_exito.html')


# ===== AUTENTICACIÓN DE ESTUDIANTES =====

def login_estudiante(request):
    """Login unificado: redirige según el rol del usuario"""
    if request.user.is_authenticated:
        return _redirigir_segun_rol(request.user)

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        usuario = authenticate(request, username=username, password=password)

        if usuario is not None:
            login(request, usuario)
            return _redirigir_segun_rol(usuario)
        else:
            messages.error(request, '❌ Usuario o contraseña incorrectos')

    return render(request, 'plataforma/login.html')


def _redirigir_segun_rol(usuario):
    """Decide a qué panel enviar al usuario según su PerfilUsuario"""
    try:
        perfil = PerfilUsuario.objects.get(user=usuario)
        if perfil.rol == 'admin':
            return redirect('panel_secretaria')
        elif perfil.rol == 'docente':
            return redirect('dashboard_docente')
        else:
            return redirect('dashboard_estudiante')
    except PerfilUsuario.DoesNotExist:
        return redirect('dashboard_estudiante')

def logout_estudiante(request):
    """Cerrar sesión"""
    logout(request)
    return redirect('index')

@login_required(login_url='login')
def dashboard_estudiante(request):
    """Panel del estudiante: sus materias matriculadas y tareas pendientes"""
    matriculas = Matricula.objects.filter(estudiante=request.user).select_related('materia', 'periodo')

    materias_ids = matriculas.values_list('materia_id', flat=True)
    tareas = Tarea.objects.filter(materia_id__in=materias_ids).select_related('materia').order_by('fecha_entrega')

    entregas_usuario = EntregaTarea.objects.filter(estudiante=request.user)
    entregas_dict = {e.tarea_id: e for e in entregas_usuario}

    tareas_con_estado = []
    for tarea in tareas:
        entrega = entregas_dict.get(tarea.id)
        tareas_con_estado.append({
            'tarea': tarea,
            'entrega': entrega,
            'estado': entrega.estado if entrega else 'Pendiente',
        })

    return render(request, 'plataforma/dashboard_estudiante.html', {
        'matriculas': matriculas,
        'tareas_con_estado': tareas_con_estado,
    })


@login_required(login_url='login')
def detalle_curso(request, matricula_id):
    """Muestra el detalle de un curso: info de la materia y sus tareas"""
    matricula = Matricula.objects.select_related('materia', 'periodo').get(
        id=matricula_id, estudiante=request.user
    )

    tareas = Tarea.objects.filter(materia=matricula.materia).order_by('fecha_entrega')
    entregas_usuario = EntregaTarea.objects.filter(estudiante=request.user, tarea__in=tareas)
    entregas_dict = {e.tarea_id: e for e in entregas_usuario}

    tareas_con_estado = []
    for tarea in tareas:
        entrega = entregas_dict.get(tarea.id)
        tareas_con_estado.append({
            'tarea': tarea,
            'entrega': entrega,
            'estado': entrega.estado if entrega else 'Pendiente',
        })

    return render(request, 'plataforma/curso_detalle.html', {
        'matricula': matricula,
        'tareas_con_estado': tareas_con_estado,
    })


@login_required(login_url='login')
def entregar_tarea(request, tarea_id):
    """Muestra una tarea y permite subir/reemplazar el archivo de entrega"""
    tarea = Tarea.objects.select_related('materia').get(id=tarea_id)

    # Seguridad: el estudiante debe estar matriculado en la materia de esta tarea
    matricula = Matricula.objects.filter(estudiante=request.user, materia=tarea.materia).first()
    if not matricula:
        messages.error(request, "❌ No tienes acceso a esta tarea")
        return redirect('dashboard_estudiante')

    entrega = EntregaTarea.objects.filter(tarea=tarea, estudiante=request.user).first()

    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        estado = 'Atrasado' if timezone.now().date() > tarea.fecha_entrega else 'Entregado'

        if entrega:
            entrega.archivo = archivo
            entrega.estado = estado
            entrega.save()
        else:
            entrega = EntregaTarea.objects.create(
                tarea=tarea, estudiante=request.user,
                archivo=archivo, estado=estado
            )

        messages.success(request, "✅ Tarea entregada correctamente")
        return redirect('entregar_tarea', tarea_id=tarea.id)

    return render(request, 'plataforma/tarea_detalle.html', {
        'tarea': tarea,
        'entrega': entrega,
        'matricula': matricula,
    })


# ===== PANEL SECRETARÍA =====

@login_required(login_url='login')
@rol_requerido('admin')
def panel_secretaria(request):
    """Dashboard principal de Secretaría"""
    usuarios_count = User.objects.count()
    carreras_count = Carrera.objects.count()
    periodos_count = Periodo.objects.count()
    materias_count = Materia.objects.count()
    estudiantes_count = PerfilEstudiante.objects.filter(activo=True).count()
    docentes_count = PerfilDocente.objects.filter(activo=True).count()

    return render(request, 'plataforma/panel_secretaria.html', {
        'usuarios_count': usuarios_count,
        'carreras_count': carreras_count,
        'periodos_count': periodos_count,
        'materias_count': materias_count,
        'estudiantes_count': estudiantes_count,
        'docentes_count': docentes_count,
    })


# ===== SECCIÓN: USUARIOS =====

@login_required(login_url='login')
@rol_requerido('admin')
def secretaria_usuarios(request):
    """Gestión de usuarios (docentes y estudiantes)"""
    rol_filtro = request.GET.get('rol', '')
    
    perfiles = PerfilUsuario.objects.select_related('user')
    if rol_filtro:
        perfiles = perfiles.filter(rol=rol_filtro)
    
    carreras = Carrera.objects.all()
    
    if request.method == 'POST':
        nombre_completo = request.POST.get('nombre_completo')
        cedula = request.POST.get('cedula')
        email = request.POST.get('email')
        rol = request.POST.get('rol')
        carrera_id = request.POST.get('carrera_id')

        # Validación previa (la cédula vive en PerfilEstudiante/PerfilDocente)
        if (PerfilEstudiante.objects.filter(cedula=cedula).exists()
                or PerfilDocente.objects.filter(cedula=cedula).exists()):
            messages.error(request, f"❌ Ya existe un usuario con cédula {cedula}")
            return redirect('secretaria_usuarios')

        if User.objects.filter(email=email).exists():
            messages.error(request, f"❌ Ya existe un usuario con email {email}")
            return redirect('secretaria_usuarios')

        try:
            with transaction.atomic():
                username = email.split('@')[0]
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    first_name=nombre_completo.split()[0],
                    last_name=' '.join(nombre_completo.split()[1:]),
                    password=make_password('Temporal123!')
                )

                carrera = Carrera.objects.get(id=carrera_id) if carrera_id else None
                PerfilUsuario.objects.create(
                    user=user,
                    rol=rol
                )

                # Crear perfil específico según rol
                if rol == 'estudiante':
                    numero_matricula = _generar_numero_matricula()
                    PerfilEstudiante.objects.create(
                        user=user,
                        cedula=cedula,
                        numero_matricula=numero_matricula,
                        carrera=carrera
                    )
                elif rol == 'docente':
                    PerfilDocente.objects.create(
                        user=user,
                        cedula=cedula,
                        carrera=carrera
                    )

                messages.success(request, f"✅ Usuario {nombre_completo} creado correctamente")
        except Exception as e:
            messages.error(request, f"❌ Error al crear usuario: {str(e)}")

        return redirect('secretaria_usuarios')

    # cedula/carrera/estado viven en PerfilEstudiante o PerfilDocente según el rol
    user_ids = [p.user_id for p in perfiles]
    estudiantes_map = {
        e.user_id: e
        for e in PerfilEstudiante.objects.select_related('carrera').filter(user_id__in=user_ids)
    }
    docentes_map = {
        d.user_id: d
        for d in PerfilDocente.objects.select_related('carrera').filter(user_id__in=user_ids)
    }

    usuarios = []
    for perfil in perfiles:
        detalle = estudiantes_map.get(perfil.user_id) or docentes_map.get(perfil.user_id)
        usuarios.append({'perfil': perfil, 'detalle': detalle})

    return render(request, 'plataforma/secretaria_usuarios.html', {
        'usuarios': usuarios,
        'carreras': carreras,
        'rol_filtro': rol_filtro,
    })


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_carga_masiva(request):
    """Carga masiva AJAX de usuarios desde Excel"""
    if not request.FILES.get('archivo_excel'):
        return JsonResponse({'ok': False, 'error': 'No se proporcionó archivo'})

    archivo = request.FILES['archivo_excel']
    try:
        wb = load_workbook(archivo)
        ws = wb.active

        password_hash = make_password('Temporal123!')
        carreras_dict = {c.id: c for c in Carrera.objects.all()}
        usernames_existentes = set(User.objects.values_list('username', flat=True))
        cedulas_existentes = (
            set(PerfilEstudiante.objects.values_list('cedula', flat=True))
            | set(PerfilDocente.objects.values_list('cedula', flat=True))
        )

        usuarios_a_crear = []
        perfiles_a_crear = []
        estudiantes_a_crear = []
        docentes_a_crear = []
        filas_validas = []
        cedulas_en_lote = set()
        usernames_en_lote = set()

        for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                cedula, nombre, email, carrera_id, rol = fila[0:5]

                if not all([cedula, nombre, email, rol]):
                    continue

                cedula = str(cedula).strip()
                username = str(email).split('@')[0]
                rol = str(rol).strip().lower()

                if rol not in ['docente', 'estudiante']:
                    continue

                if cedula in cedulas_existentes or cedula in cedulas_en_lote:
                    continue

                if username in usernames_existentes or username in usernames_en_lote:
                    continue

                carrera = carreras_dict.get(int(carrera_id)) if carrera_id else None

                cedulas_en_lote.add(cedula)
                usernames_en_lote.add(username)

                usuario = User(
                    username=username,
                    email=email,
                    first_name=nombre.split()[0] if nombre else '',
                    last_name=' '.join(nombre.split()[1:]) if len(nombre.split()) > 1 else '',
                    password=password_hash
                )
                usuarios_a_crear.append(usuario)
                filas_validas.append((usuario, cedula, rol, carrera))

            except Exception:
                continue

        with transaction.atomic():
            User.objects.bulk_create(usuarios_a_crear)

            usuarios_creados = {u.username: u for u in User.objects.filter(username__in=[u.username for u in usuarios_a_crear])}

            for usuario, cedula, rol, carrera in filas_validas:
                user_obj = usuarios_creados.get(usuario.username)
                if user_obj:
                    perfiles_a_crear.append(PerfilUsuario(
                        user=user_obj,
                        rol=rol
                    ))

                    if rol == 'estudiante':
                        numero_matricula = _generar_numero_matricula()
                        estudiantes_a_crear.append(PerfilEstudiante(
                            user=user_obj,
                            cedula=cedula,
                            numero_matricula=numero_matricula,
                            carrera=carrera
                        ))
                    elif rol == 'docente':
                        docentes_a_crear.append(PerfilDocente(
                            user=user_obj,
                            cedula=cedula,
                            carrera=carrera
                        ))

            PerfilUsuario.objects.bulk_create(perfiles_a_crear)
            if estudiantes_a_crear:
                PerfilEstudiante.objects.bulk_create(estudiantes_a_crear)
            if docentes_a_crear:
                PerfilDocente.objects.bulk_create(docentes_a_crear)

        return JsonResponse({'ok': True, 'creados': len(perfiles_a_crear), 'omitidos': 0})

    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


def _generar_numero_matricula():
    """Genera un número de matrícula único: YYYYMMDD-XXXX"""
    fecha = datetime.now().strftime('%Y%m%d')
    contador = PerfilEstudiante.objects.filter(numero_matricula__startswith=fecha).count()
    return f"{fecha}-{str(contador + 1).zfill(4)}"


# ===== SECCIÓN: CARRERAS =====

@login_required(login_url='login')
@rol_requerido('admin')
def secretaria_carreras(request):
    """Gestión de carreras"""
    carreras = Carrera.objects.all()

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        codigo = request.POST.get('codigo')
        descripcion = request.POST.get('descripcion', '')

        if Carrera.objects.filter(codigo=codigo).exists():
            messages.error(request, f"❌ Ya existe una carrera con código {codigo}")
            return redirect('secretaria_carreras')

        try:
            Carrera.objects.create(
                nombre=nombre,
                codigo=codigo,
                descripcion=descripcion
            )
            messages.success(request, f"✅ Carrera {nombre} creada correctamente")
        except Exception as e:
            messages.error(request, f"❌ Error: {str(e)}")

        return redirect('secretaria_carreras')

    return render(request, 'plataforma/secretaria_carreras.html', {
        'carreras': carreras,
    })


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_carreras_editar(request, carrera_id):
    """Editar carrera (AJAX JSON)"""
    carrera = get_object_or_404(Carrera, id=carrera_id)

    try:
        carrera.nombre = request.POST.get('nombre', carrera.nombre)
        carrera.codigo = request.POST.get('codigo', carrera.codigo)
        carrera.descripcion = request.POST.get('descripcion', carrera.descripcion)
        carrera.save()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_carreras_eliminar(request, carrera_id):
    """Eliminar carrera (AJAX JSON) - con protección"""
    carrera = get_object_or_404(Carrera, id=carrera_id)

    # Protección: no eliminar si tiene relaciones
    if Materia.objects.filter(carrera=carrera).exists():
        return JsonResponse({'ok': False, 'error': 'Esta carrera tiene materias asociadas'})
    if PerfilEstudiante.objects.filter(carrera=carrera).exists():
        return JsonResponse({'ok': False, 'error': 'Esta carrera tiene estudiantes'})

    try:
        carrera.delete()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_carreras_carga_masiva(request):
    """Carga masiva de carreras desde Excel"""
    if not request.FILES.get('archivo_excel'):
        return JsonResponse({'ok': False, 'error': 'No se proporcionó archivo'})

    archivo = request.FILES['archivo_excel']
    try:
        wb = load_workbook(archivo)
        ws = wb.active

        codigos_existentes = set(Carrera.objects.values_list('codigo', flat=True))
        carreras_a_crear = []
        codigos_en_lote = set()

        for fila in ws.iter_rows(min_row=2, values_only=True):
            try:
                nombre, codigo = fila[0], fila[1]
                descripcion = fila[2] if len(fila) > 2 else ''

                if not nombre or not codigo:
                    continue

                codigo = str(codigo).strip()

                if codigo in codigos_existentes or codigo in codigos_en_lote:
                    continue

                codigos_en_lote.add(codigo)
                carreras_a_crear.append(Carrera(
                    nombre=nombre,
                    codigo=codigo,
                    descripcion=descripcion or ''
                ))

            except Exception:
                continue

        Carrera.objects.bulk_create(carreras_a_crear)
        return JsonResponse({'ok': True, 'creados': len(carreras_a_crear)})

    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


# ===== SECCIÓN: PERÍODOS =====

@login_required(login_url='login')
@rol_requerido('admin')
def secretaria_periodos(request):
    """Gestión de períodos académicos"""
    periodos = Periodo.objects.all().order_by('-fecha_inicio')

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')

        try:
            Periodo.objects.create(
                nombre=nombre,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin
            )
            messages.success(request, f"✅ Período {nombre} creado correctamente")
        except Exception as e:
            messages.error(request, f"❌ Error: {str(e)}")

        return redirect('secretaria_periodos')

    return render(request, 'plataforma/secretaria_periodos.html', {
        'periodos': periodos,
    })


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_periodos_editar(request, periodo_id):
    """Editar período (AJAX JSON)"""
    periodo = get_object_or_404(Periodo, id=periodo_id)

    try:
        periodo.nombre = request.POST.get('nombre', periodo.nombre)
        periodo.fecha_inicio = request.POST.get('fecha_inicio', periodo.fecha_inicio)
        periodo.fecha_fin = request.POST.get('fecha_fin', periodo.fecha_fin)
        periodo.save()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_periodos_eliminar(request, periodo_id):
    """Eliminar período (AJAX JSON)"""
    periodo = get_object_or_404(Periodo, id=periodo_id)

    try:
        periodo.delete()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_periodos_carga_masiva(request):
    """Carga masiva de períodos desde Excel"""
    if not request.FILES.get('archivo_excel'):
        return JsonResponse({'ok': False, 'error': 'No se proporcionó archivo'})

    archivo = request.FILES['archivo_excel']
    try:
        wb = load_workbook(archivo)
        ws = wb.active

        periodos_a_crear = []

        for fila in ws.iter_rows(min_row=2, values_only=True):
            try:
                nombre, fecha_inicio, fecha_fin = fila[0], fila[1], fila[2]

                if not all([nombre, fecha_inicio, fecha_fin]):
                    continue

                periodos_a_crear.append(Periodo(
                    nombre=nombre,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin
                ))

            except Exception:
                continue

        Periodo.objects.bulk_create(periodos_a_crear)
        return JsonResponse({'ok': True, 'creados': len(periodos_a_crear)})

    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


# ===== SECCIÓN: MATERIAS =====

@login_required(login_url='login')
@rol_requerido('admin')
def secretaria_materias(request):
    """Gestión de materias"""
    materias = Materia.objects.select_related('carrera').all()
    carreras = Carrera.objects.all()

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        codigo = request.POST.get('codigo')
        carrera_id = request.POST.get('carrera_id')
        creditos = request.POST.get('creditos', 4)
        profesor = request.POST.get('profesor', '')

        if Materia.objects.filter(codigo=codigo).exists():
            messages.error(request, f"❌ Ya existe una materia con código {codigo}")
            return redirect('secretaria_materias')

        try:
            carrera = Carrera.objects.get(id=carrera_id)
            Materia.objects.create(
                nombre=nombre,
                codigo=codigo,
                carrera=carrera,
                creditos=int(creditos),
                profesor=profesor
            )
            messages.success(request, f"✅ Materia {nombre} creada correctamente")
        except Exception as e:
            messages.error(request, f"❌ Error: {str(e)}")

        return redirect('secretaria_materias')

    return render(request, 'plataforma/secretaria_materias.html', {
        'materias': materias,
        'carreras': carreras,
    })


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_materias_editar(request, materia_id):
    """Editar materia (AJAX JSON)"""
    materia = get_object_or_404(Materia, id=materia_id)

    try:
        materia.nombre = request.POST.get('nombre', materia.nombre)
        materia.codigo = request.POST.get('codigo', materia.codigo)
        materia.creditos = request.POST.get('creditos', materia.creditos)
        materia.profesor = request.POST.get('profesor', materia.profesor)
        materia.save()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_materias_eliminar(request, materia_id):
    """Eliminar materia (AJAX JSON)"""
    materia = get_object_or_404(Materia, id=materia_id)

    try:
        materia.delete()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_materias_carga_masiva(request):
    """Carga masiva de materias desde Excel"""
    if not request.FILES.get('archivo_excel'):
        return JsonResponse({'ok': False, 'error': 'No se proporcionó archivo'})

    archivo = request.FILES['archivo_excel']
    try:
        wb = load_workbook(archivo)
        ws = wb.active

        carreras_dict = {c.id: c for c in Carrera.objects.all()}
        codigos_existentes = set(Materia.objects.values_list('codigo', flat=True))
        materias_a_crear = []
        codigos_en_lote = set()

        for fila in ws.iter_rows(min_row=2, values_only=True):
            try:
                nombre, codigo, carrera_id, creditos = fila[0], fila[1], fila[2], fila[3]
                profesor = fila[4] if len(fila) > 4 else ''

                if not all([nombre, codigo, carrera_id]):
                    continue

                codigo = str(codigo).strip()

                if codigo in codigos_existentes or codigo in codigos_en_lote:
                    continue

                carrera = carreras_dict.get(int(carrera_id))
                if not carrera:
                    continue

                codigos_en_lote.add(codigo)
                materias_a_crear.append(Materia(
                    nombre=nombre,
                    codigo=codigo,
                    carrera=carrera,
                    creditos=int(creditos) if creditos else 4,
                    profesor=profesor or ''
                ))

            except Exception:
                continue

        Materia.objects.bulk_create(materias_a_crear)
        return JsonResponse({'ok': True, 'creados': len(materias_a_crear)})

    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


# ===== SECCIÓN: MATRÍCULAS =====

@login_required(login_url='login')
@rol_requerido('admin')
def secretaria_matriculas(request):
    """Gestión de matrículas"""
    matriculas = Matricula.objects.select_related('estudiante__user', 'materia', 'periodo').all()
    estudiantes = PerfilEstudiante.objects.select_related('user').filter(activo=True)
    materias = Materia.objects.all()
    periodos = Periodo.objects.all()

    if request.method == 'POST':
        estudiante_id = request.POST.get('estudiante_id')
        materia_id = request.POST.get('materia_id')
        periodo_id = request.POST.get('periodo_id')

        try:
            estudiante = PerfilEstudiante.objects.get(id=estudiante_id)
            materia = Materia.objects.get(id=materia_id)
            periodo = Periodo.objects.get(id=periodo_id)

            if Matricula.objects.filter(estudiante=estudiante, materia=materia, periodo=periodo).exists():
                messages.error(request, "❌ Esta matrícula ya existe")
                return redirect('secretaria_matriculas')

            Matricula.objects.create(
                estudiante=estudiante,
                materia=materia,
                periodo=periodo
            )
            messages.success(request, f"✅ Matrícula creada correctamente")
        except Exception as e:
            messages.error(request, f"❌ Error: {str(e)}")

        return redirect('secretaria_matriculas')

    return render(request, 'plataforma/secretaria_matriculas.html', {
        'matriculas': matriculas,
        'estudiantes': estudiantes,
        'materias': materias,
        'periodos': periodos,
    })


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_matriculas_editar(request, matricula_id):
    """Editar matrícula (AJAX JSON)"""
    matricula = get_object_or_404(Matricula, id=matricula_id)

    try:
        matricula.nota_parcial_1 = request.POST.get('nota_parcial_1', matricula.nota_parcial_1)
        matricula.nota_parcial_2 = request.POST.get('nota_parcial_2', matricula.nota_parcial_2)
        matricula.nota_final = request.POST.get('nota_final', matricula.nota_final)
        matricula.asistencia = request.POST.get('asistencia', matricula.asistencia)
        matricula.estado = request.POST.get('estado', matricula.estado)
        matricula.save()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_matriculas_eliminar(request, matricula_id):
    """Eliminar matrícula (AJAX JSON)"""
    matricula = get_object_or_404(Matricula, id=matricula_id)

    try:
        matricula.delete()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_matriculas_carga_masiva(request):
    """Carga masiva de matrículas desde Excel"""
    if not request.FILES.get('archivo_excel'):
        return JsonResponse({'ok': False, 'error': 'No se proporcionó archivo'})

    archivo = request.FILES['archivo_excel']
    try:
        wb = load_workbook(archivo)
        ws = wb.active

        estudiantes_dict = {e.cedula: e for e in PerfilEstudiante.objects.all()}
        materias_dict = {m.codigo: m for m in Materia.objects.all()}
        periodos_dict = {p.nombre: p for p in Periodo.objects.all()}
        matriculas_a_crear = []

        for fila in ws.iter_rows(min_row=2, values_only=True):
            try:
                cedula_estudiante, codigo_materia, periodo_nombre = fila[0], fila[1], fila[2]

                if not all([cedula_estudiante, codigo_materia, periodo_nombre]):
                    continue

                estudiante = estudiantes_dict.get(str(cedula_estudiante).strip())
                materia = materias_dict.get(str(codigo_materia).strip())
                periodo = periodos_dict.get(str(periodo_nombre).strip())

                if not all([estudiante, materia, periodo]):
                    continue

                if not Matricula.objects.filter(estudiante=estudiante, materia=materia, periodo=periodo).exists():
                    matriculas_a_crear.append(Matricula(
                        estudiante=estudiante,
                        materia=materia,
                        periodo=periodo
                    ))

            except Exception:
                continue

        Matricula.objects.bulk_create(matriculas_a_crear)
        return JsonResponse({'ok': True, 'creados': len(matriculas_a_crear)})

    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


# ===== SECCIÓN: CERTIFICADOS =====

@login_required(login_url='login')
@rol_requerido('admin')
def secretaria_certificados(request):
    """Gestión de certificados"""
    tipo_filtro = request.GET.get('tipo', '')
    
    certificados = Certificado.objects.select_related('estudiante__user').all()
    if tipo_filtro:
        certificados = certificados.filter(tipo=tipo_filtro)
    
    estudiantes = PerfilEstudiante.objects.select_related('user').filter(activo=True)

    if request.method == 'POST':
        estudiante_id = request.POST.get('estudiante_id')
        tipo = request.POST.get('tipo')
        fecha_emision = request.POST.get('fecha_emision')

        try:
            estudiante = PerfilEstudiante.objects.get(id=estudiante_id)
            numero_tramite = _generar_numero_tramite()
            Certificado.objects.create(
                estudiante=estudiante,
                tipo=tipo,
                fecha_emision=fecha_emision,
                numero_tramite=numero_tramite
            )
            messages.success(request, f"✅ Certificado generado correctamente")
        except Exception as e:
            messages.error(request, f"❌ Error: {str(e)}")

        return redirect('secretaria_certificados')

    return render(request, 'plataforma/secretaria_certificados.html', {
        'certificados': certificados,
        'estudiantes': estudiantes,
        'tipo_filtro': tipo_filtro,
    })


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_certificados_eliminar(request, certificado_id):
    """Eliminar certificado (AJAX JSON)"""
    certificado = get_object_or_404(Certificado, id=certificado_id)

    try:
        certificado.delete()
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


@login_required(login_url='login')
@rol_requerido('admin')
@require_http_methods(["POST"])
def secretaria_certificados_carga_masiva(request):
    """Carga masiva de certificados desde Excel"""
    if not request.FILES.get('archivo_excel'):
        return JsonResponse({'ok': False, 'error': 'No se proporcionó archivo'})

    archivo = request.FILES['archivo_excel']
    try:
        wb = load_workbook(archivo)
        ws = wb.active

        estudiantes_dict = {e.cedula: e for e in PerfilEstudiante.objects.all()}
        certificados_a_crear = []

        for fila in ws.iter_rows(min_row=2, values_only=True):
            try:
                cedula_estudiante, tipo, fecha_emision = fila[0], fila[1], fila[2]

                if not all([cedula_estudiante, tipo]):
                    continue

                estudiante = estudiantes_dict.get(str(cedula_estudiante).strip())
                if not estudiante:
                    continue

                numero_tramite = _generar_numero_tramite()
                certificados_a_crear.append(Certificado(
                    estudiante=estudiante,
                    tipo=tipo,
                    fecha_emision=fecha_emision or timezone.now().date(),
                    numero_tramite=numero_tramite
                ))

            except Exception:
                continue

        Certificado.objects.bulk_create(certificados_a_crear)
        return JsonResponse({'ok': True, 'creados': len(certificados_a_crear)})

    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)})


def _generar_numero_tramite():
    """Genera un número de trámite único"""
    fecha = datetime.now().strftime('%Y%m%d')
    contador = Certificado.objects.filter(numero_tramite__startswith=fecha).count()
    return f"{fecha}-{str(contador + 1).zfill(5)}"


@login_required(login_url='login')
@rol_requerido('admin')
def descargar_certificado(request, certificado_id):
    """Descargar certificado (PDF)"""
    certificado = get_object_or_404(Certificado, id=certificado_id)
    # TODO: implementar generación de PDF
    messages.info(request, "📄 Función de descarga PDF en construcción")
    return redirect('secretaria_certificados')


# ===== IMPORTAR DATOS DESDE EXCEL (legacy - mantener compatibilidad) =====

def _importar_estudiantes_desde_excel(ws):
    """Procesa un Excel de estudiantes. Columnas: Nombre, Apellido, Email, Cédula, Nº Matrícula, Carrera ID"""
    password_hash = make_password('Temporal123!')
    carreras_dict = {c.id: c for c in Carrera.objects.all()}
    usernames_existentes = set(User.objects.values_list('username', flat=True))
    cedulas_existentes = set(PerfilEstudiante.objects.values_list('cedula', flat=True))
    matriculas_existentes = set(PerfilEstudiante.objects.values_list('numero_matricula', flat=True))

    usuarios_a_crear = []
    filas_validas = []
    errores = []
    usernames_en_lote = set()
    cedulas_en_lote = set()
    matriculas_en_lote = set()

    for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            nombre, apellido, email, cedula, numero_matricula, carrera_id = fila[0:6]

            if not all([nombre, apellido, email, cedula, numero_matricula]):
                errores.append(f"Fila {fila_num}: Faltan datos obligatorios")
                continue

            username = str(email).split('@')[0]
            cedula = str(cedula).strip()
            numero_matricula = str(numero_matricula).strip()

            if username in usernames_existentes or username in usernames_en_lote:
                errores.append(f"Fila {fila_num}: El usuario {username} ya existe")
                continue
            if cedula in cedulas_existentes or cedula in cedulas_en_lote:
                errores.append(f"Fila {fila_num}: La cédula {cedula} ya existe")
                continue
            if numero_matricula in matriculas_existentes or numero_matricula in matriculas_en_lote:
                errores.append(f"Fila {fila_num}: La matrícula {numero_matricula} ya existe")
                continue

            carrera = carreras_dict.get(int(carrera_id))
            if not carrera:
                errores.append(f"Fila {fila_num}: La carrera con ID {carrera_id} no existe")
                continue

            usernames_en_lote.add(username)
            cedulas_en_lote.add(cedula)
            matriculas_en_lote.add(numero_matricula)

            usuarios_a_crear.append(User(
                username=username, email=email,
                first_name=nombre, last_name=apellido,
                password=password_hash
            ))
            filas_validas.append((username, cedula, numero_matricula, carrera))

        except Exception as e:
            errores.append(f"Fila {fila_num}: {str(e)}")

    with transaction.atomic():
        User.objects.bulk_create(usuarios_a_crear)

        usuarios_creados = {
            u.username: u for u in User.objects.filter(username__in=[f[0] for f in filas_validas])
        }

        perfiles_a_crear = []
        roles_a_crear = []
        for username, cedula, numero_matricula, carrera in filas_validas:
            usuario_obj = usuarios_creados.get(username)
            if usuario_obj:
                perfiles_a_crear.append(PerfilEstudiante(
                    user=usuario_obj, carrera=carrera,
                    cedula=cedula, numero_matricula=numero_matricula
                ))
                roles_a_crear.append(PerfilUsuario(user=usuario_obj, rol='estudiante'))

        PerfilEstudiante.objects.bulk_create(perfiles_a_crear)
        PerfilUsuario.objects.bulk_create(roles_a_crear)

    return len(perfiles_a_crear), errores


def _importar_carreras_desde_excel(ws):
    """Procesa un Excel de carreras. Columnas: Nombre, Código, Descripción, Créditos totales"""
    codigos_existentes = set(Carrera.objects.values_list('codigo', flat=True))
    carreras_a_crear = []
    errores = []
    codigos_en_lote = set()

    for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            nombre, codigo = fila[0], fila[1]
            descripcion = fila[2] if len(fila) > 2 and fila[2] else ''
            creditos_totales = fila[3] if len(fila) > 3 and fila[3] else 120

            if not all([nombre, codigo]):
                errores.append(f"Fila {fila_num}: Faltan datos obligatorios")
                continue

            if codigo in codigos_existentes or codigo in codigos_en_lote:
                errores.append(f"Fila {fila_num}: El código de carrera {codigo} ya existe")
                continue

            codigos_en_lote.add(codigo)
            carreras_a_crear.append(Carrera(
                nombre=nombre, codigo=codigo,
                descripcion=descripcion, creditos_totales=int(creditos_totales)
            ))

        except Exception as e:
            errores.append(f"Fila {fila_num}: {str(e)}")

    Carrera.objects.bulk_create(carreras_a_crear)
    return len(carreras_a_crear), errores


def _importar_docentes_desde_excel(ws):
    """Procesa un Excel de docentes. Columnas: Nombre, Apellido, Email, Cédula, Título académico, Carrera ID"""
    password_hash = make_password('Temporal123!')
    carreras_dict = {c.id: c for c in Carrera.objects.all()}
    usernames_existentes = set(User.objects.values_list('username', flat=True))
    cedulas_existentes = set(PerfilDocente.objects.values_list('cedula', flat=True))

    usuarios_a_crear = []
    filas_validas = []
    errores = []
    usernames_en_lote = set()
    cedulas_en_lote = set()

    for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            nombre, apellido, email, cedula, titulo, carrera_id = fila[0:6]

            if not all([nombre, apellido, email, cedula]):
                errores.append(f"Fila {fila_num}: Faltan datos obligatorios")
                continue

            username = str(email).split('@')[0]
            cedula = str(cedula).strip()

            if username in usernames_existentes or username in usernames_en_lote:
                errores.append(f"Fila {fila_num}: El usuario {username} ya existe")
                continue
            if cedula in cedulas_existentes or cedula in cedulas_en_lote:
                errores.append(f"Fila {fila_num}: La cédula {cedula} ya existe")
                continue

            carrera = carreras_dict.get(int(carrera_id)) if carrera_id else None
            if carrera_id and not carrera:
                errores.append(f"Fila {fila_num}: La carrera con ID {carrera_id} no existe")
                continue

            usernames_en_lote.add(username)
            cedulas_en_lote.add(cedula)

            usuarios_a_crear.append(User(
                username=username, email=email,
                first_name=nombre, last_name=apellido,
                password=password_hash
            ))
            filas_validas.append((username, cedula, titulo or '', carrera))

        except Exception as e:
            errores.append(f"Fila {fila_num}: {str(e)}")

    with transaction.atomic():
        User.objects.bulk_create(usuarios_a_crear)

        usuarios_creados = {
            u.username: u for u in User.objects.filter(username__in=[f[0] for f in filas_validas])
        }

        perfiles_a_crear = []
        roles_a_crear = []
        for username, cedula, titulo, carrera in filas_validas:
            usuario_obj = usuarios_creados.get(username)
            if usuario_obj:
                perfiles_a_crear.append(PerfilDocente(
                    user=usuario_obj, carrera=carrera,
                    cedula=cedula, titulo_academico=titulo
                ))
                roles_a_crear.append(PerfilUsuario(user=usuario_obj, rol='docente'))

        PerfilDocente.objects.bulk_create(perfiles_a_crear)
        PerfilUsuario.objects.bulk_create(roles_a_crear)

    return len(perfiles_a_crear), errores


def _importar_materias_desde_excel(ws):
    """Procesa un Excel de materias. Columnas: Nombre, Código, Carrera ID, Créditos, Profesor"""
    carreras_dict = {c.id: c for c in Carrera.objects.all()}
    codigos_existentes = set(Materia.objects.values_list('codigo', flat=True))
    materias_a_crear = []
    errores = []
    codigos_en_lote = set()

    for fila_num, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            nombre, codigo, carrera_id = fila[0], fila[1], fila[2]
            creditos = fila[3] if len(fila) > 3 and fila[3] else 4
            profesor = fila[4] if len(fila) > 4 and fila[4] else ''

            if not all([nombre, codigo, carrera_id]):
                errores.append(f"Fila {fila_num}: Faltan datos obligatorios")
                continue

            if codigo in codigos_existentes or codigo in codigos_en_lote:
                errores.append(f"Fila {fila_num}: El código de materia {codigo} ya existe")
                continue

            carrera = carreras_dict.get(int(carrera_id))
            if not carrera:
                errores.append(f"Fila {fila_num}: La carrera con ID {carrera_id} no existe")
                continue

            codigos_en_lote.add(codigo)
            materias_a_crear.append(Materia(
                nombre=nombre, codigo=codigo, carrera=carrera,
                creditos=int(creditos), profesor=profesor
            ))

        except Exception as e:
            errores.append(f"Fila {fila_num}: {str(e)}")

    Materia.objects.bulk_create(materias_a_crear)
    return len(materias_a_crear), errores


@staff_member_required
def importar_estudiantes(request):
    """Importación masiva de datos (Estudiantes, Carreras o Materias) desde Excel"""
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        tipo = request.POST.get('tipo_importacion', 'estudiantes')

        try:
            wb = load_workbook(archivo)
            ws = wb.active

            if tipo == 'estudiantes':
                contador, errores = _importar_estudiantes_desde_excel(ws)
                etiqueta = 'estudiantes'
            elif tipo == 'carreras':
                contador, errores = _importar_carreras_desde_excel(ws)
                etiqueta = 'carreras'
            elif tipo == 'materias':
                contador, errores = _importar_materias_desde_excel(ws)
                etiqueta = 'materias'
            elif tipo == 'docentes':
                contador, errores = _importar_docentes_desde_excel(ws)
                etiqueta = 'docentes'
            else:
                messages.error(request, "❌ Tipo de importación no reconocido")
                return redirect('importar_estudiantes')

            if contador > 0:
                messages.success(request, f"✅ {contador} {etiqueta} importados correctamente")

            if errores:
                for error in errores[:10]:
                    messages.warning(request, error)
                if len(errores) > 10:
                    messages.info(request, f"... y {len(errores) - 10} errores más")

            return redirect('importar_estudiantes')

        except Exception as e:
            messages.error(request, f"❌ Error al importar: {str(e)}")

    carreras = Carrera.objects.all()
    return render(request, 'plataforma/importar_estudiantes.html', {'carreras': carreras})


# ===== PANEL DOCENTE =====
def _perfil_docente_o_none(user):
    """Devuelve el PerfilDocente del usuario, o None si no lo tiene configurado."""
    return PerfilDocente.objects.select_related('carrera').filter(user=user).first()


@login_required(login_url='login')
@rol_requerido('docente')
def dashboard_docente(request):
    """Panel del docente: sus materias del período activo con su carga de trabajo"""
    perfil = _perfil_docente_o_none(request.user)
    if not perfil:
        messages.error(request, "❌ Tu perfil de docente no está configurado. Contacta a secretaría.")
        return redirect('index')

    periodo = Periodo.objects.filter(activo=True).order_by('-fecha_inicio').first()
    materias = perfil.materias.select_related('carrera').order_by('codigo')
    materias_ids = list(materias.values_list('id', flat=True))

    # Conteos por materia, en una consulta por métrica en vez de una por materia
    matriculas_qs = Matricula.objects.filter(materia_id__in=materias_ids)
    if periodo:
        matriculas_qs = matriculas_qs.filter(periodo=periodo)
    estudiantes_por_materia = {
        fila['materia_id']: fila['total']
        for fila in matriculas_qs.values('materia_id').annotate(total=Count('id'))
    }

    tareas_por_materia = {
        fila['materia_id']: fila['total']
        for fila in Tarea.objects.filter(materia_id__in=materias_ids)
                                 .values('materia_id').annotate(total=Count('id'))
    }

    pendientes_por_materia = {
        fila['tarea__materia_id']: fila['total']
        for fila in EntregaTarea.objects.filter(
            tarea__materia_id__in=materias_ids, calificacion__isnull=True
        ).values('tarea__materia_id').annotate(total=Count('id'))
    }

    planificaciones = {}
    if periodo:
        planificaciones = {
            plan.materia_id: plan
            for plan in Planificacion.objects.filter(materia_id__in=materias_ids, periodo=periodo)
        }

    tarjetas = []
    for materia in materias:
        tarjetas.append({
            'materia': materia,
            'estudiantes': estudiantes_por_materia.get(materia.id, 0),
            'tareas': tareas_por_materia.get(materia.id, 0),
            'por_calificar': pendientes_por_materia.get(materia.id, 0),
            'planificacion': planificaciones.get(materia.id),
        })

    return render(request, 'plataforma/dashboard_docente.html', {
        'perfil': perfil,
        'periodo': periodo,
        'materias': tarjetas,
        'total_materias': len(tarjetas),
        'total_estudiantes': sum(t['estudiantes'] for t in tarjetas),
        'total_por_calificar': sum(t['por_calificar'] for t in tarjetas),
    })


def _materia_del_docente(user, materia_id):
    """
    Devuelve (perfil_docente, materia) solo si la materia pertenece a ese docente.
    Devuelve (perfil, None) o (None, None) cuando no debe tener acceso.
    """
    perfil = _perfil_docente_o_none(user)
    if not perfil:
        return None, None
    materia = Materia.objects.select_related('carrera').filter(id=materia_id, docente=perfil).first()
    return perfil, materia


@login_required(login_url='login')
@rol_requerido('docente')
def docente_materia_tareas(request, materia_id):
    """Tareas de una materia del docente: listarlas y crear nuevas"""
    perfil, materia = _materia_del_docente(request.user, materia_id)
    if not perfil:
        messages.error(request, "❌ Tu perfil de docente no está configurado. Contacta a secretaría.")
        return redirect('index')
    if not materia:
        messages.error(request, "❌ Esa materia no está asignada a ti")
        return redirect('dashboard_docente')

    if request.method == 'POST':
        titulo = (request.POST.get('titulo') or '').strip()
        descripcion = (request.POST.get('descripcion') or '').strip()
        fecha_entrega = request.POST.get('fecha_entrega')

        if not titulo or not fecha_entrega:
            messages.error(request, "❌ El título y la fecha de entrega son obligatorios")
            return redirect('docente_materia_tareas', materia_id=materia.id)

        try:
            fecha = datetime.strptime(fecha_entrega, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "❌ Fecha de entrega inválida")
            return redirect('docente_materia_tareas', materia_id=materia.id)

        tarea = Tarea(
            materia=materia,
            titulo=titulo,
            descripcion=descripcion,
            fecha_entrega=fecha,
        )
        if request.FILES.get('pdf_guia'):
            tarea.pdf_guia = request.FILES['pdf_guia']
        tarea.save()

        messages.success(request, f"✅ Tarea «{tarea.titulo}» creada")
        return redirect('docente_materia_tareas', materia_id=materia.id)

    tareas = Tarea.objects.filter(materia=materia).order_by('-fecha_entrega')
    total_matriculados = Matricula.objects.filter(materia=materia).count()

    entregas_por_tarea = {
        fila['tarea_id']: fila['total']
        for fila in EntregaTarea.objects.filter(tarea__materia=materia)
                                        .values('tarea_id').annotate(total=Count('id'))
    }
    calificadas_por_tarea = {
        fila['tarea_id']: fila['total']
        for fila in EntregaTarea.objects.filter(tarea__materia=materia, calificacion__isnull=False)
                                        .values('tarea_id').annotate(total=Count('id'))
    }

    filas = []
    for tarea in tareas:
        entregadas = entregas_por_tarea.get(tarea.id, 0)
        filas.append({
            'tarea': tarea,
            'entregadas': entregadas,
            'calificadas': calificadas_por_tarea.get(tarea.id, 0),
            'por_calificar': entregadas - calificadas_por_tarea.get(tarea.id, 0),
        })

    return render(request, 'plataforma/docente_materia_tareas.html', {
        'perfil': perfil,
        'materia': materia,
        'tareas': filas,
        'total_matriculados': total_matriculados,
    })


@login_required(login_url='login')
@rol_requerido('docente')
def docente_tarea_entregas(request, tarea_id):
    """Entregas de una tarea: quién entregó, quién no, y la nota puesta"""
    perfil = _perfil_docente_o_none(request.user)
    if not perfil:
        messages.error(request, "❌ Tu perfil de docente no está configurado. Contacta a secretaría.")
        return redirect('index')

    tarea = Tarea.objects.select_related('materia', 'materia__carrera').filter(
        id=tarea_id, materia__docente=perfil
    ).first()
    if not tarea:
        messages.error(request, "❌ Esa tarea no pertenece a una materia tuya")
        return redirect('dashboard_docente')

    if request.method == 'POST':
        entrega_id = request.POST.get('entrega_id')
        nota_texto = (request.POST.get('calificacion') or '').strip()
        observacion = (request.POST.get('observacion') or '').strip()

        entrega = EntregaTarea.objects.filter(id=entrega_id, tarea=tarea).first()
        if not entrega:
            messages.error(request, "❌ Entrega no encontrada")
            return redirect('docente_tarea_entregas', tarea_id=tarea.id)

        try:
            nota = float(nota_texto)
        except ValueError:
            messages.error(request, "❌ La calificación debe ser un número")
            return redirect('docente_tarea_entregas', tarea_id=tarea.id)

        if not 0 <= nota <= 100:
            messages.error(request, "❌ La calificación debe estar entre 0 y 100")
            return redirect('docente_tarea_entregas', tarea_id=tarea.id)

        entrega.calificacion = nota
        entrega.observacion = observacion
        entrega.estado = 'Calificado'
        entrega.save()

        messages.success(request, f"✅ Calificación guardada para {entrega.estudiante.get_full_name()}")
        return redirect('docente_tarea_entregas', tarea_id=tarea.id)

    entregas = EntregaTarea.objects.filter(tarea=tarea).select_related('estudiante')
    entregas_por_estudiante = {e.estudiante_id: e for e in entregas}

    matriculas = Matricula.objects.filter(materia=tarea.materia).select_related('estudiante')

    filas = []
    for matricula in matriculas:
        filas.append({
            'estudiante': matricula.estudiante,
            'entrega': entregas_por_estudiante.get(matricula.estudiante_id),
        })
    filas.sort(key=lambda f: (f['estudiante'].last_name, f['estudiante'].first_name))

    return render(request, 'plataforma/docente_tarea_entregas.html', {
        'perfil': perfil,
        'tarea': tarea,
        'materia': tarea.materia,
        'filas': filas,
        'total_entregadas': sum(1 for f in filas if f['entrega']),
        'total_por_calificar': sum(1 for f in filas if f['entrega'] and f['entrega'].calificacion is None),
    })


@login_required(login_url='login')
@rol_requerido('docente')
def docente_materia_estudiantes(request, materia_id):
    """Lista de estudiantes matriculados en una materia del docente"""
    perfil, materia = _materia_del_docente(request.user, materia_id)
    if not perfil:
        messages.error(request, "❌ Tu perfil de docente no está configurado. Contacta a secretaría.")
        return redirect('index')
    if not materia:
        messages.error(request, "❌ Esa materia no está asignada a ti")
        return redirect('dashboard_docente')

    periodo_id = request.GET.get('periodo', '')
    matriculas = Matricula.objects.filter(materia=materia).select_related('estudiante', 'periodo')
    if periodo_id:
        matriculas = matriculas.filter(periodo_id=periodo_id)

    estudiantes_ids = [m.estudiante_id for m in matriculas]

    perfiles = {
        pe.user_id: pe
        for pe in PerfilEstudiante.objects.filter(user_id__in=estudiantes_ids)
    }

    total_tareas = Tarea.objects.filter(materia=materia).count()
    entregas_por_estudiante = {
        fila['estudiante_id']: fila['total']
        for fila in EntregaTarea.objects.filter(tarea__materia=materia, estudiante_id__in=estudiantes_ids)
                                        .values('estudiante_id').annotate(total=Count('id'))
    }
    promedios = {
        fila['estudiante_id']: fila['promedio']
        for fila in EntregaTarea.objects.filter(
            tarea__materia=materia, estudiante_id__in=estudiantes_ids, calificacion__isnull=False
        ).values('estudiante_id').annotate(promedio=Avg('calificacion'))
    }

    filas = []
    for matricula in matriculas:
        filas.append({
            'matricula': matricula,
            'estudiante': matricula.estudiante,
            'perfil': perfiles.get(matricula.estudiante_id),
            'entregas': entregas_por_estudiante.get(matricula.estudiante_id, 0),
            'promedio_entregas': promedios.get(matricula.estudiante_id),
        })
    filas.sort(key=lambda f: (f['estudiante'].last_name, f['estudiante'].first_name))

    periodos = Periodo.objects.filter(matriculas__materia=materia).distinct().order_by('-fecha_inicio')

    return render(request, 'plataforma/docente_materia_estudiantes.html', {
        'perfil': perfil,
        'materia': materia,
        'filas': filas,
        'total_tareas': total_tareas,
        'periodos': periodos,
        'periodo_filtro': periodo_id,
    })


@login_required(login_url='login')
@rol_requerido('docente')
def docente_materia_planificacion(request, materia_id):
    """Parciales de la planificación activa de una materia, con acceso para calificarlos"""
    perfil, materia = _materia_del_docente(request.user, materia_id)
    if not perfil:
        messages.error(request, "❌ Tu perfil de docente no está configurado. Contacta a secretaría.")
        return redirect('index')
    if not materia:
        messages.error(request, "❌ Esa materia no está asignada a ti")
        return redirect('dashboard_docente')

    periodo = Periodo.objects.filter(activo=True).order_by('-fecha_inicio').first()
    planificacion = None
    if periodo:
        planificacion = Planificacion.objects.filter(materia=materia, periodo=periodo).first()

    parciales = []
    if planificacion:
        parciales_qs = planificacion.parciales.prefetch_related('unidades__actividades').order_by('numero')
        for parcial in parciales_qs:
            unidades = list(parcial.unidades.all())
            parciales.append({
                'parcial': parcial,
                'total_unidades': len(unidades),
                'total_actividades': sum(len(unidad.actividades.all()) for unidad in unidades),
            })

    return render(request, 'plataforma/docente_materia_planificacion.html', {
        'perfil': perfil,
        'materia': materia,
        'periodo': periodo,
        'planificacion': planificacion,
        'parciales': parciales,
    })


def _promedio(valores):
    """Promedio simple, o None si no hay notas."""
    return sum(valores) / len(valores) if valores else None


def _calcular_nota_parcial(parcial, notas_por_categoria):
    """
    Aplica la fórmula del parcial sobre los promedios de cada categoría.

        nota = (prom_formativa  × peso_formativa / 100)
             + (prom_practica   × peso_practica  / 100)
             + (nota_examen     × peso_examen    / 100)

    Una categoría sin notas cuenta como 0 y la nota se marca incompleta, para que
    el docente vea que aún falta evaluar algo en vez de recibir un número que
    parece definitivo.
    """
    prom_formativa = _promedio(notas_por_categoria.get('formativa', []))
    prom_practica = _promedio(notas_por_categoria.get('practica', []))
    nota_examen = _promedio(notas_por_categoria.get('examen', []))

    nota = (
        (prom_formativa or 0) * parcial.peso_formativa / 100
        + (prom_practica or 0) * parcial.peso_practica / 100
        + (nota_examen or 0) * parcial.peso_examen / 100
    )

    faltantes = [
        etiqueta
        for valor, peso, etiqueta in (
            (prom_formativa, parcial.peso_formativa, 'formativa'),
            (prom_practica, parcial.peso_practica, 'práctica'),
            (nota_examen, parcial.peso_examen, 'examen'),
        )
        if valor is None and peso > 0
    ]

    return {
        'prom_formativa': prom_formativa,
        'prom_practica': prom_practica,
        'nota_examen': nota_examen,
        'nota': round(nota, 2),
        'faltantes': faltantes,
        'completa': not faltantes,
    }


@login_required(login_url='login')
@rol_requerido('docente')
def docente_calificar_parcial(request, materia_id, parcial_id):
    """Calcula y guarda la nota de un parcial aplicando los pesos por categoría"""
    perfil, materia = _materia_del_docente(request.user, materia_id)
    if not perfil:
        messages.error(request, "❌ Tu perfil de docente no está configurado. Contacta a secretaría.")
        return redirect('index')
    if not materia:
        messages.error(request, "❌ Esa materia no está asignada a ti")
        return redirect('dashboard_docente')

    parcial = Parcial.objects.select_related('planificacion', 'planificacion__periodo').filter(
        id=parcial_id, planificacion__materia=materia
    ).first()
    if not parcial:
        messages.error(request, "❌ Ese parcial no pertenece a la planificación de esta materia")
        return redirect('dashboard_docente')

    tipo_calificacion = f"Parcial {parcial.numero}"
    tipos_validos = [t[0] for t in Calificacion.TIPOS]

    periodo = parcial.planificacion.periodo
    matriculas = Matricula.objects.filter(materia=materia, periodo=periodo).select_related('estudiante')

    # Entregas calificadas de tareas que cuelgan de este parcial
    entregas = EntregaTarea.objects.filter(
        tarea__materia=materia,
        tarea__actividad__unidad__parcial=parcial,
        calificacion__isnull=False,
    ).select_related('tarea', 'tarea__actividad')

    notas = {}
    for entrega in entregas:
        por_categoria = notas.setdefault(entrega.estudiante_id, {})
        por_categoria.setdefault(entrega.tarea.actividad.categoria, []).append(entrega.calificacion)

    filas = []
    for matricula in matriculas:
        calculo = _calcular_nota_parcial(parcial, notas.get(matricula.estudiante_id, {}))
        calculo['matricula'] = matricula
        calculo['estudiante'] = matricula.estudiante
        filas.append(calculo)
    filas.sort(key=lambda f: (f['estudiante'].last_name, f['estudiante'].first_name))

    if request.method == 'POST':
        if tipo_calificacion not in tipos_validos:
            messages.error(
                request,
                f"❌ No se puede guardar «{tipo_calificacion}»: Calificacion solo admite "
                f"{', '.join(tipos_validos)}."
            )
            return redirect('docente_calificar_parcial', materia_id=materia.id, parcial_id=parcial.id)

        solo_completas = request.POST.get('solo_completas') == '1'
        guardadas = 0

        with transaction.atomic():
            for fila in filas:
                if solo_completas and not fila['completa']:
                    continue

                Calificacion.objects.update_or_create(
                    matricula=fila['matricula'],
                    tipo=tipo_calificacion,
                    defaults={'valor': fila['nota']},
                )
                guardadas += 1

                # La nota final de la materia es el promedio de los parciales guardados
                valores = list(
                    Calificacion.objects.filter(
                        matricula=fila['matricula'], tipo__startswith='Parcial '
                    ).values_list('valor', flat=True)
                )
                if valores:
                    fila['matricula'].nota_final = round(sum(valores) / len(valores), 2)
                    fila['matricula'].save(update_fields=['nota_final'])

        messages.success(request, f"✅ {tipo_calificacion}: {guardadas} nota(s) guardadas")
        return redirect('docente_calificar_parcial', materia_id=materia.id, parcial_id=parcial.id)

    return render(request, 'plataforma/docente_calificar_parcial.html', {
        'perfil': perfil,
        'materia': materia,
        'parcial': parcial,
        'periodo': periodo,
        'filas': filas,
        'tipo_calificacion': tipo_calificacion,
        'tipo_soportado': tipo_calificacion in tipos_validos,
        'total_incompletas': sum(1 for f in filas if not f['completa']),
    })
